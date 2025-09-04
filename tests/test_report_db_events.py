import asyncio
import json
import time
from datetime import datetime
from io import BytesIO

import fakeredis
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import config as config_mod
from modules.events_store import RedisStore
from routers import reports


def _setup_context(tmp_path):
    redis_client = fakeredis.FakeRedis()
    cfg = {
        "track_objects": ["person", "vehicle"],
        "branding": {},
        "logo_url": "",
        "count_classes": ["person", "vehicle"],
    }
    config_mod.set_config(cfg)
    reports.init_context(cfg, {}, redis_client, str(tmp_path), [])
    app = FastAPI()
    app.include_router(reports.router)
    reports.require_roles = lambda request, roles: True
    client = TestClient(app)
    store = RedisStore(redis_client)
    return client, store


def test_persist_and_report(tmp_path):
    client, store = _setup_context(tmp_path)
    now = int(time.time())
    events = [
        (now - 3, "in", "person"),
        (now - 2, "out", "person"),
        (now - 1, "in", "vehicle"),
        (now, "out", "vehicle"),
    ]
    for idx, (ts, direction, label) in enumerate(events, 1):
        store.persist_event(
            ts_utc=ts,
            ts_local=datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M"),
            camera_id=1,
            camera_name="cam",
            track_id=idx,
            direction=direction,
            label=label,
            image_path="",
            thumb_path="",
        )
    assert len(store.fetch_events(now - 10, now + 1)) == 4
    assert store.count_events(["person"], "in", now - 10, now + 1) == 1
    start = datetime.fromtimestamp(now - 60).strftime("%Y-%m-%d %H:%M")
    end = datetime.fromtimestamp(now + 60).strftime("%Y-%m-%d %H:%M")
    params_person = {
        "start": start,
        "end": end,
        "type": "person",
        "view": "table",
        "rows": 10,
        "cam_id": "",
        "label": "person",
        "cursor": 0,
    }
    res = client.get("/report_data", params=params_person)
    assert res.status_code == 200
    data = res.json()
    assert len(data["rows"]) == 2
    t0 = events[0][0]
    assert data["rows"][1]["time"] == datetime.fromtimestamp(t0).strftime("%Y-%m-%d %H:%M")
    params_vehicle = {**params_person, "label": "vehicle", "type": "vehicle"}
    res_v = client.get("/report_data", params=params_vehicle)
    assert len(res_v.json()["rows"]) == 2
    res_export = client.get("/report/export", params=params_person)
    wb = load_workbook(BytesIO(res_export.content))
    assert wb.active.max_row == 3
